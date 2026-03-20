import base64
import os
import io
from datetime import datetime

try:
    import openai
except ImportError:
    openai = None

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
except ImportError:
    A4 = colors = ParagraphStyle = mm = SimpleDocTemplate = Paragraph = Spacer = HRFlowable = Table = TableStyle = TA_LEFT = TA_CENTER = None

try:
    from pdf2image import convert_from_bytes
    PDF2IMAGE_AVAILABLE = True
except ImportError:
    convert_from_bytes = None
    PDF2IMAGE_AVAILABLE = False

if load_dotenv:
    load_dotenv()


def pdf_to_image_bytes(pdf_file, page=0, dpi=200):
    """Convert first page of a PDF to PNG bytes. Returns (success, bytes|error_str)."""
    if not PDF2IMAGE_AVAILABLE:
        return False, "pdf2image is not installed."
    try:
        pdf_file.seek(0)
        pdf_bytes = pdf_file.read()
        pages = convert_from_bytes(pdf_bytes, dpi=dpi, first_page=page+1, last_page=page+1)
        if not pages:
            return False, "Could not extract any pages from the PDF."
        buf = io.BytesIO()
        pages[0].save(buf, format="PNG")
        buf.seek(0)
        return True, buf.read()
    except Exception as e:
        return False, str(e)

PROXY_URL = os.getenv("PROXY_URL", "https://web-production-a87eb.up.railway.app/v1")

if openai:
    client = openai.OpenAI(
        api_key=os.getenv("OPENAI_API_KEY", "draft-ai-proxy"),
        base_url=PROXY_URL
    )
else:
    client = None


def _require_openai():
    if client is None:
        raise RuntimeError("OpenAI dependency is unavailable. Install the 'openai' package and redeploy.")


def _require_reportlab():
    if SimpleDocTemplate is None:
        raise RuntimeError("PDF generation dependency is unavailable. Install 'reportlab' and redeploy.")

# ── BASE FORMATTING RULES (shared by all prompts) ──
FORMAT_RULES = """
STRICT FORMATTING RULES — follow these exactly, no exceptions:
- NEVER use markdown headers like #, ##, ###
- NEVER use asterisks for bold like **text**
- Use plain numbered lists like: 1. item, 2. item
- Use plain bullet lists starting with: - item
- Use CAPS for section titles followed by a colon, like: DIMENSIONS:
- Keep answers clear, direct, and structured
- Separate sections with a blank line
- No filler phrases like "Great question" or "Certainly"
"""

SYSTEM_PROMPT = f"""You are an expert mechanical engineer who reads and interprets engineering drawings with precision.
{FORMAT_RULES}
Your capabilities:
- Extract all dimensions, tolerances, and annotations accurately
- Identify drawing type (assembly, detail, section, isometric etc.)
- Read and explain GD&T symbols
- Extract title block information
- Flag design concerns, missing info, or standard violations
- Answer follow-up questions using conversation context"""


GDT_PROMPT = f"""You are a senior mechanical engineer and GD&T specialist certified in ASME Y14.5.
{FORMAT_RULES}
When analyzing GD&T symbols, always respond in this exact structure:

SYMBOLS DETECTED:
- List every GD&T symbol found with its location on the drawing

DETAILED EXPLANATION:
- For each symbol: name, what it controls, the tolerance value, and the datum reference

CORRECTNESS ASSESSMENT:
- Is each symbol applied correctly per ASME Y14.5?
- Are datum references logical and complete?
- Are tolerance values realistic for the feature?

ISSUES FOUND:
- List any missing, incorrect, or conflicting GD&T callouts
- If none, say: No issues detected

RECOMMENDATIONS:
- Suggest any improvements to the GD&T scheme"""


DESIGN_CONCERN_PROMPT = f"""You are a senior mechanical design engineer with 20 years of experience reviewing engineering drawings for production readiness.
{FORMAT_RULES}
Analyze the drawing thoroughly and respond in this exact structure:

SEVERITY LEGEND:
- CRITICAL: Will cause part failure or cannot be manufactured
- WARNING: May cause issues in manufacturing or assembly
- INFO: Minor improvement suggestions

DESIGN CONCERNS:
- Number each concern with its severity level like: 1. [CRITICAL] Missing datum reference on feature...
- Be specific about location on the drawing

MISSING INFORMATION:
- List any required callouts, tolerances, or notes that are absent

STANDARD VIOLATIONS:
- List any deviations from ASME/ISO drawing standards

MANUFACTURABILITY ISSUES:
- Features that are difficult or impossible to machine as drawn

OVERALL ASSESSMENT:
- Rate the drawing: PRODUCTION READY / NEEDS REVISION / MAJOR REWORK REQUIRED
- Give a one-line summary"""


MATERIAL_PROMPT = f"""You are a materials engineer and manufacturing consultant specializing in mechanical component design.
{FORMAT_RULES}
Analyze the drawing carefully and respond in this exact structure:

SPECIFIED MATERIAL:
- What material is explicitly called out in the drawing (or "Not specified")

ANALYSIS OF REQUIREMENTS:
- Loading conditions visible from the drawing (stress concentrations, thin walls, bearing surfaces etc.)
- Environmental factors to consider
- Surface finish requirements

PRIMARY RECOMMENDATION:
- Material name and grade (e.g. Aluminum 6061-T6)
- Why it suits this component
- Typical yield strength, density, machinability rating

ALTERNATIVE OPTIONS:
1. [Material] — [reason, trade-offs]
2. [Material] — [reason, trade-offs]
3. [Material] — [reason, trade-offs]

MATERIALS TO AVOID:
- List materials that would be unsuitable and why

HEAT TREATMENT / SURFACE TREATMENT:
- Recommended post-processing for the specified or recommended material"""


MANUFACTURING_PROMPT = f"""You are a manufacturing engineer with expertise in CNC machining, casting, forging, additive manufacturing, and production optimization.
{FORMAT_RULES}
Analyze the drawing and respond in this exact structure:

COMPONENT OVERVIEW:
- Type of part, estimated complexity, and key features driving manufacturing decisions

PRIMARY MANUFACTURING METHOD:
- Recommended process (e.g. CNC Turning + Milling)
- Why it suits this geometry
- Estimated number of setups required

OPERATION SEQUENCE:
1. List each machining operation in order
2. Include tool types where relevant
3. Note critical features requiring precision

ALTERNATIVE METHODS:
- Method 1: [name] — suitable if [condition], trade-off: [cost/time/quality]
- Method 2: [name] — suitable if [condition], trade-off: [cost/time/quality]

CRITICAL FEATURES:
- Features requiring special attention, fixtures, or tooling

TOLERANCING REVIEW:
- Are the tolerances achievable with standard equipment?
- Flag any tight tolerances that require grinding/EDM/etc.

ESTIMATED PRODUCTION NOTES:
- Suitable for: [low volume / medium volume / high volume]
- Key cost drivers on this part"""


def _call_vision_api(image_file, system_prompt, user_message, max_tokens=1400):
    """Internal helper — single place where we call the API. No debug prints."""
    _require_openai()
    base64_image = base64.b64encode(image_file.read()).decode("utf-8")
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/jpeg;base64,{base64_image}", "detail": "high"}
                    },
                    {"type": "text", "text": user_message}
                ]
            }
        ],
        max_tokens=max_tokens
    )
    return response.choices[0].message.content


def _call_vision_api_with_history(image_file, system_prompt, question, chat_history, max_tokens=1400):
    """Vision API call with conversation history."""
    _require_openai()
    base64_image = base64.b64encode(image_file.read()).decode("utf-8")
    messages = [{"role": "system", "content": system_prompt}]
    for msg in chat_history:
        messages.append(msg)
    messages.append({
        "role": "user",
        "content": [
            {
                "type": "image_url",
                "image_url": {"url": f"data:image/jpeg;base64,{base64_image}", "detail": "high"}
            },
            {"type": "text", "text": question}
        ]
    })
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=messages,
        max_tokens=max_tokens
    )
    return response.choices[0].message.content


# ── PUBLIC FUNCTIONS ──

def analyze_drawing(image_file, question, chat_history=[]):
    """General Q&A analysis with chat history."""
    return _call_vision_api_with_history(image_file, SYSTEM_PROMPT, question, chat_history)


def analyze_gdt(image_file):
    """Deep GD&T symbol detection and explanation."""
    return _call_vision_api(
        image_file,
        GDT_PROMPT,
        "Perform a complete GD&T analysis of this engineering drawing. Identify every symbol, explain each one, and assess correctness.",
        max_tokens=1600
    )


def analyze_design_concerns(image_file):
    """Detect design issues, missing info, and standard violations."""
    return _call_vision_api(
        image_file,
        DESIGN_CONCERN_PROMPT,
        "Perform a thorough design review of this engineering drawing. Identify all concerns, issues, and violations.",
        max_tokens=1600
    )


def analyze_material(image_file):
    """Material analysis and recommendations."""
    return _call_vision_api(
        image_file,
        MATERIAL_PROMPT,
        "Analyze this engineering drawing and provide a complete material recommendation with alternatives and reasoning.",
        max_tokens=1400
    )


def analyze_manufacturing(image_file):
    """Manufacturing method suggestions and operation sequence."""
    return _call_vision_api(
        image_file,
        MANUFACTURING_PROMPT,
        "Analyze this engineering drawing and recommend the best manufacturing methods, operation sequence, and production notes.",
        max_tokens=1600
    )


def detect_dimensions(image_file):
    """OCR + Vision dimension detection — returns structured JSON."""
    return _call_vision_api(
        image_file,
        """You are an expert mechanical engineer specializing in engineering drawing interpretation.
Extract ALL dimensions from the drawing and return ONLY a valid JSON object — no explanation, no markdown, no backticks.

Return this exact JSON structure:
{
  "dimensions": [
    {
      "label": "Outer Diameter",
      "value": "5.75",
      "unit": "inches",
      "tolerance": "±0.005",
      "type": "diameter",
      "location": "top view, center"
    }
  ],
  "summary": "12 dimensions detected. Units: inches. General tolerance: ±0.01"
}

Types can be: diameter, radius, length, width, height, depth, angle, thread, chamfer, fillet, other
If tolerance is not specified, use the drawing's general tolerance or write "per general tolerance".
If a value is unclear, use your best reading and add "(approx)" to the value.
Return ONLY the JSON. Nothing else.""",
        "Extract every dimension from this engineering drawing and return structured JSON.",
        max_tokens=1800
    )



def extract_title_block(image_file):
    """Extract title block key-value pairs."""
    return _call_vision_api(
        image_file,
        """You are an expert at reading engineering drawing title blocks.
Extract all title block information and return it as plain key-value pairs.
Format exactly like this (no markdown, no asterisks, no headers):

Part Name: [value or Not specified]
Part Number: [value or Not specified]
Material: [value or Not specified]
Scale: [value or Not specified]
Drawing Number: [value or Not specified]
Revision: [value or Not specified]
Drawn By: [value or Not specified]
Checked By: [value or Not specified]
Date: [value or Not specified]
Company: [value or Not specified]
Tolerance: [value or Not specified]
Surface Finish: [value or Not specified]
Units: [value or Not specified]

Only include fields that are visible or can be inferred. Keep values short and factual.""",
        "Extract all title block information from this engineering drawing.",
        max_tokens=600
    )



# ══════════════════════════════════════════════════════════════════
# FEATURE 1 — TOLERANCE STACK-UP ANALYSIS
# ══════════════════════════════════════════════════════════════════

TOLERANCE_STACKUP_PROMPT = f"""You are a senior dimensional analyst and metrology engineer specializing in tolerance stack-up analysis.
{FORMAT_RULES}
Analyze all dimensions and tolerances visible in the drawing and respond in this exact structure:

DIMENSIONAL CHAINS IDENTIFIED:
- List each chain of dimensions that form a functional loop
- Format: Chain [letter]: [dim1] + [dim2] - [dim3] = [result feature]

WORST-CASE ANALYSIS:
- For each chain: Nominal = X | Max = X | Min = X | Total variation = ±X
- Flag any chain where worst-case exceeds an acceptable limit

RSS (STATISTICAL) ANALYSIS:
- Root Sum Square estimate for each chain
- RSS result is typically 60-70% of worst-case — note where this matters

CRITICAL FITS & GAPS:
- List all clearance fits, interference fits, or gap closures
- State whether each fit is ACCEPTABLE or AT RISK under worst-case

CONFLICT ALERTS:
- Any tolerance that makes a fit impossible under worst-case conditions
- Tolerance pairs that are contradictory or over-constrained

RECOMMENDATIONS:
- Which specific tolerances to tighten (and by how much)
- Which tolerances can be loosened to reduce cost
- Where GD&T would better control the critical dimension chains

OVERALL RISK LEVEL: LOW / MEDIUM / HIGH
- One sentence explaining the biggest risk in this drawing"""


def analyze_tolerance_stackup(image_file):
    """Tolerance stack-up: worst-case and RSS analysis of dimensional chains."""
    return _call_vision_api(
        image_file,
        TOLERANCE_STACKUP_PROMPT,
        "Perform a complete tolerance stack-up analysis on this engineering drawing. Identify all dimensional chains, compute worst-case and RSS, and flag any at-risk fits.",
        max_tokens=1800
    )


# ══════════════════════════════════════════════════════════════════
# FEATURE 2 — MANUFACTURABILITY SCORE
# ══════════════════════════════════════════════════════════════════

MANUFACTURABILITY_PROMPT = f"""You are a manufacturing engineer who evaluates engineering drawings for ease and cost of manufacture.
{FORMAT_RULES}
Score the drawing across 6 categories (each out of the listed max) and respond in this exact structure:

MANUFACTURABILITY SCORE: [total] / 100

CATEGORY SCORES:
- Geometry Complexity     [X / 20]: [one-line reason]
- Tolerance Achievability [X / 20]: [one-line reason]
- Surface Finish Specs    [X / 15]: [one-line reason]
- Drawing Completeness    [X / 15]: [one-line reason]
- Material Machinability  [X / 15]: [one-line reason]
- Setup & Fixturing Ease  [X / 15]: [one-line reason]

WHAT'S HELPING THE SCORE:
- List 3-5 specific features or decisions that make this part easy to manufacture

WHAT'S HURTING THE SCORE:
- List 3-5 specific features, tolerances, or missing callouts dragging the score down

DIFFICULTY RATING: EASY / MODERATE / COMPLEX / EXPERT-ONLY

TOP 3 DESIGN-FOR-MANUFACTURE IMPROVEMENTS:
1. [Most impactful change — be specific]
2. [Second improvement]
3. [Third improvement]

SUITABLE PRODUCTION VOLUMES:
- Best suited for: [prototype / low-volume / medium-volume / high-volume] — reason"""


def analyze_manufacturability_score(image_file):
    """Score the drawing 0-100 for manufacturability with full category breakdown."""
    return _call_vision_api(
        image_file,
        MANUFACTURABILITY_PROMPT,
        "Score the manufacturability of this engineering drawing from 0 to 100. Break down each category score and provide specific improvement suggestions.",
        max_tokens=1600
    )


# ══════════════════════════════════════════════════════════════════
# FEATURE 3 — COST ESTIMATION
# ══════════════════════════════════════════════════════════════════

COST_ESTIMATION_PROMPT = f"""You are a manufacturing cost estimator with expertise in machined, cast, sheet-metal, and fabricated components.
{FORMAT_RULES}
Analyze the drawing and provide a detailed cost estimate in this exact structure:

PART OVERVIEW:
- Estimated bounding box / envelope dimensions
- Estimated finished weight (material + assumed density)
- Primary manufacturing process assumed
- Material assumed (or read from drawing)

COST BREAKDOWN — LOW VOLUME (1–10 pieces):

MATERIAL COST:
- Raw stock type and estimated buy size
- Estimated material cost per part: $X – $X

MACHINING / PROCESS COST:
- Setup time: ~X hrs × $X/hr = $X (one-time per batch)
- Cycle time per part: ~X hrs × $X/hr = $X
- Number of setups required: X
- Subtotal machining: $X – $X per part

FINISHING & TREATMENT COST:
- Required operations (deburring, anodize, plating, heat treat, etc.): $X – $X

TOTAL UNIT COST ESTIMATES:
- 1–10 pcs:   $[low] – $[high] per part
- 100 pcs:    $[low] – $[high] per part
- 1,000 pcs:  $[low] – $[high] per part

TOP COST DRIVERS:
1. [Feature or spec driving the most cost]
2. [Second driver]
3. [Third driver]

COST REDUCTION OPPORTUNITIES:
- 3 specific drawing changes that would reduce cost without compromising function

DISCLAIMER: Rough order-of-magnitude estimates only. Actual supplier quotes will vary."""


def estimate_cost(image_file):
    """Rough cost estimation broken down by material, machining, and finishing."""
    return _call_vision_api(
        image_file,
        COST_ESTIMATION_PROMPT,
        "Estimate the manufacturing cost of this part. Break down material, machining, and finishing costs across low, medium, and high volumes.",
        max_tokens=1800
    )


# ══════════════════════════════════════════════════════════════════
# FEATURE 4 — MISSING DIMENSION DETECTION
# ══════════════════════════════════════════════════════════════════

MISSING_DIMENSIONS_PROMPT = f"""You are a senior drawing checker with deep expertise in ASME Y14.5 and ISO GPS standards.
{FORMAT_RULES}
Carefully inspect every visible feature and respond in this exact structure:

MISSING DIMENSIONS:
- List every feature that lacks a required size or location dimension
- Format: [Feature name / location on drawing] — missing: [what is needed]
- If nothing is missing: All features appear fully dimensioned

AMBIGUOUS DIMENSIONS:
- Dimensions that could be interpreted multiple ways
- Dimensions with unclear datum references
- Dimensions whose origin or direction is not clear from the views

REDUNDANT / CONFLICTING DIMENSIONS:
- Over-constrained features where dimensions close a loop incorrectly
- Duplicate dimensions that could conflict if they disagree

MISSING TOLERANCES:
- Features that have a dimension but no associated tolerance where one is functionally required
- Missing general tolerance note (if none exists on the drawing)

MISSING ANNOTATIONS:
- Surface finish symbols absent from functional surfaces
- Thread specifications incomplete (missing pitch, class, depth)
- Material or finish treatment callouts absent
- Weld symbols, heat treat notes, or other required specifications not present

MISSING OR INADEQUATE VIEWS:
- Features that cannot be fully understood from the current views
- Recommended additional views or sections that would complete the drawing

COMPLETENESS SCORE: [X / 100]

SUMMARY:
- One short paragraph describing the overall completeness of this drawing and the most critical gaps"""


def detect_missing_dimensions(image_file):
    """Check the drawing for missing dimensions, tolerances, annotations, and views."""
    return _call_vision_api(
        image_file,
        MISSING_DIMENSIONS_PROMPT,
        "Check this engineering drawing thoroughly for missing dimensions, tolerances, annotations, and views. List everything that is absent or ambiguous.",
        max_tokens=1800
    )


# ══════════════════════════════════════════════════════════════════
# FEATURE 5 — DRAWING REVISION COMPARISON
# ══════════════════════════════════════════════════════════════════

REVISION_COMPARISON_PROMPT = f"""You are an expert drawing reviewer comparing two revisions of an engineering drawing.
{FORMAT_RULES}
You will receive two images: Revision A (older) and Revision B (newer).
Compare them carefully and respond in this exact structure:

REVISION SUMMARY:
- One paragraph overview: what kind of changes were made and how significant they are

DIMENSIONAL CHANGES:
- List every dimension that changed
- Format: [Feature] — Rev A: [value] → Rev B: [value]
- If none detected: No dimensional changes found

TOLERANCE CHANGES:
- List every tolerance that changed (value, type, or datum reference)
- If none: No tolerance changes found

GEOMETRY & FEATURE CHANGES:
- New features added in Rev B
- Features removed from Rev A
- Features that changed shape, position, count, or size

ANNOTATION & NOTE CHANGES:
- Changed surface finish callouts
- Changed material or treatment specifications
- Changed general notes or BOM entries

TITLE BLOCK CHANGES:
- Revision letter, date, drawn by, approval — what changed

GD&T CHANGES:
- Added, removed, or modified GD&T callouts

RISK ASSESSMENT:
- [CRITICAL] Changes that affect form, fit, or function
- [WARNING]  Changes that affect manufacturing process or cost
- [INFO]     Administrative or cosmetic changes only

RECOMMENDATION: APPROVE / NEEDS REVIEW / REJECT
- One sentence justifying the recommendation"""


def compare_revisions(image_file_a, image_file_b):
    """Compare two drawing revisions side-by-side and list all changes."""
    b64_a = base64.b64encode(image_file_a.read()).decode("utf-8")
    b64_b = base64.b64encode(image_file_b.read()).decode("utf-8")

    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": REVISION_COMPARISON_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text",      "text": "REVISION A — the older drawing:"},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_a}", "detail": "high"}},
                    {"type": "text",      "text": "REVISION B — the newer drawing:"},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64_b}", "detail": "high"}},
                    {"type": "text",      "text": "Compare these two drawing revisions carefully and report every change you find."},
                ],
            }
        ],
        max_tokens=2000
    )
    return response.choices[0].message.content



# ==================================================================
# BATCH ANALYSIS
# ==================================================================

BATCH_PROMPT = f"""You are a senior mechanical engineer reviewing an engineering drawing for a batch production report.
{FORMAT_RULES}
Analyze this drawing and return ONLY a valid JSON object — no explanation, no markdown, no backticks.

Return exactly this structure:
{{
  "drawing_name": "inferred from title block or use filename",
  "part_number": "from title block or Not specified",
  "drawing_type": "Detail / Assembly / Section / Schematic / Other",
  "status": "Production Ready / Needs Revision / Major Rework Required",
  "manufacturability_score": 85,
  "estimated_cost_usd": "45-65",
  "complexity": "Low / Medium / High / Expert",
  "critical_issues": ["issue 1", "issue 2"],
  "warnings": ["warning 1", "warning 2"],
  "missing_dimensions": true,
  "has_gdt": true,
  "material_specified": false,
  "tolerance_risk": "Low / Medium / High",
  "recommended_process": "CNC Turning + Milling",
  "summary": "One sentence summary of this drawing"
}}

Be precise. Return ONLY the JSON. Nothing else."""


def batch_analyze_drawing(image_file, filename="drawing"):
    """Analyze a single drawing and return structured JSON for batch report."""
    try:
        result = _call_vision_api(
            image_file,
            BATCH_PROMPT,
            f"Analyze this engineering drawing (filename: {filename}) and return the structured JSON report.",
            max_tokens=800
        )
        import json, re
        clean = result.strip()
        if "```" in clean:
            clean = re.sub(r'```[a-z]*', '', clean).replace("```", "").strip()
        return json.loads(clean)
    except Exception as e:
        return {
            "drawing_name": filename,
            "part_number": "—",
            "drawing_type": "Unknown",
            "status": "Analysis Failed",
            "manufacturability_score": 0,
            "estimated_cost_usd": "—",
            "complexity": "—",
            "critical_issues": [str(e)],
            "warnings": [],
            "missing_dimensions": False,
            "has_gdt": False,
            "material_specified": False,
            "tolerance_risk": "—",
            "recommended_process": "—",
            "summary": "Analysis could not be completed for this drawing."
        }


def generate_batch_excel(results):
    """Generate an Excel report from batch analysis results."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Analysis Report"

    # ── Colour palette ──
    ORANGE      = "F97316"
    DARK        = "0D0D0D"
    HEADER_BG   = "1A1A1A"
    ROW_A       = "F9F9F9"
    ROW_B       = "FFFFFF"
    RED_BG      = "FEE2E2"
    YELLOW_BG   = "FEF9C3"
    GREEN_BG    = "DCFCE7"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = f"Draft AI — Batch Analysis Report   |   {datetime.now().strftime('%d %B %Y, %I:%M %p')}   |   {len(results)} drawings"
    title_cell.font      = Font(name="Calibri", size=13, bold=True, color=ORANGE)
    title_cell.fill      = PatternFill("solid", fgColor=DARK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Column headers ──
    headers = [
        "#", "Drawing Name", "Part Number", "Type", "Status",
        "Mfg. Score", "Est. Cost (USD)", "Complexity",
        "Tolerance Risk", "Missing Dims", "Has GD&T",
        "Material Specified", "Process", "Summary"
    ]
    col_widths = [4, 28, 16, 14, 20, 12, 16, 12, 14, 12, 10, 16, 24, 40]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 20

    # ── Data rows ──
    for ri, r in enumerate(results, 1):
        row = ri + 2
        bg = ROW_A if ri % 2 == 0 else ROW_B

        # Status colour
        status = r.get("status", "")
        if "Major" in status or "Failed" in status:
            status_bg = RED_BG
        elif "Revision" in status:
            status_bg = YELLOW_BG
        else:
            status_bg = GREEN_BG

        # Score colour
        score = r.get("manufacturability_score", 0)
        if isinstance(score, (int, float)):
            if score >= 75:   score_bg = GREEN_BG
            elif score >= 50: score_bg = YELLOW_BG
            else:             score_bg = RED_BG
        else:
            score_bg = bg

        vals = [
            ri,
            r.get("drawing_name", "—"),
            r.get("part_number", "—"),
            r.get("drawing_type", "—"),
            status,
            score,
            r.get("estimated_cost_usd", "—"),
            r.get("complexity", "—"),
            r.get("tolerance_risk", "—"),
            "Yes" if r.get("missing_dimensions") else "No",
            "Yes" if r.get("has_gdt") else "No",
            "Yes" if r.get("material_specified") else "No",
            r.get("recommended_process", "—"),
            r.get("summary", "—"),
        ]

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 14))
            cell.border    = border

            # Apply special bg
            if ci == 5:   cell.fill = PatternFill("solid", fgColor=status_bg)
            elif ci == 6: cell.fill = PatternFill("solid", fgColor=score_bg)
            else:         cell.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row].height = 18

    # ── Issues sheet ──
    ws2 = wb.create_sheet("Issues & Warnings")
    ws2["A1"].value = "Draft AI — Critical Issues & Warnings"
    ws2["A1"].font  = Font(name="Calibri", size=13, bold=True, color=ORANGE)
    ws2["A1"].fill  = PatternFill("solid", fgColor=DARK)
    ws2.merge_cells("A1:E1")
    ws2.row_dimensions[1].height = 28

    issue_headers = ["#", "Drawing Name", "Type", "Severity", "Issue / Warning"]
    issue_widths  = [4, 28, 10, 12, 70]
    for ci, (h, w) in enumerate(zip(issue_headers, issue_widths), 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws2.column_dimensions[get_column_letter(ci)].width = w

    irow = 3
    for ri, r in enumerate(results, 1):
        name = r.get("drawing_name", "—")
        for issue in r.get("critical_issues", []):
            for ci, val in enumerate([ri, name, "CRITICAL", "Critical", issue], 1):
                cell = ws2.cell(row=irow, column=ci, value=val)
                cell.font   = Font(name="Calibri", size=10)
                cell.fill   = PatternFill("solid", fgColor=RED_BG)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=(ci==5))
            ws2.row_dimensions[irow].height = 18
            irow += 1
        for warn in r.get("warnings", []):
            for ci, val in enumerate([ri, name, "WARNING", "Warning", warn], 1):
                cell = ws2.cell(row=irow, column=ci, value=val)
                cell.font   = Font(name="Calibri", size=10)
                cell.fill   = PatternFill("solid", fgColor=YELLOW_BG)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=(ci==5))
            ws2.row_dimensions[irow].height = 18
            irow += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_batch_pdf(results):
    """Generate a PDF summary report from batch analysis results."""
    _require_reportlab()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=18*mm, leftMargin=18*mm,
        topMargin=18*mm, bottomMargin=18*mm
    )

    title_s  = ParagraphStyle('T',  fontSize=18, fontName='Helvetica-Bold', textColor=colors.HexColor('#f97316'), spaceAfter=3)
    sub_s    = ParagraphStyle('S',  fontSize=9,  fontName='Helvetica',      textColor=colors.HexColor('#888888'), spaceAfter=2)
    h2_s     = ParagraphStyle('H2', fontSize=11, fontName='Helvetica-Bold', textColor=colors.HexColor('#111111'), spaceBefore=10, spaceAfter=5)
    body_s   = ParagraphStyle('B',  fontSize=9,  fontName='Helvetica',      textColor=colors.HexColor('#333333'), leading=14, spaceAfter=3)
    foot_s   = ParagraphStyle('F',  fontSize=7,  fontName='Helvetica',      textColor=colors.HexColor('#aaaaaa'), alignment=TA_CENTER)

    story = []
    # Header block — no overlap
    batch_hdr_data = [[
        Paragraph('<font color="#f97316"><b>Draft AI</b></font>',
            ParagraphStyle('BHL', fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#f97316'), leading=26)),
        Paragraph(datetime.now().strftime("%d %B %Y, %I:%M %p"),
            ParagraphStyle('BHR', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#aaaaaa'), leading=12, alignment=2)),
    ]]
    batch_hdr_tbl = Table(batch_hdr_data, colWidths=[80*mm, None])
    batch_hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),0),('TOPPADDING',(0,0),(-1,-1),0)]))
    story.append(batch_hdr_tbl)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(f"Batch Analysis Report  —  {len(results)} drawings", sub_s))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#f97316'), spaceAfter=5*mm))

    # Summary stats
    total   = len(results)
    ready   = sum(1 for r in results if "Ready" in r.get("status",""))
    needs   = sum(1 for r in results if "Revision" in r.get("status",""))
    rework  = sum(1 for r in results if "Major" in r.get("status","") or "Failed" in r.get("status",""))
    scores  = [r.get("manufacturability_score",0) for r in results if isinstance(r.get("manufacturability_score"),int)]
    avg_sc  = round(sum(scores)/len(scores)) if scores else "—"

    story.append(Paragraph("SUMMARY", h2_s))
    summary_data = [
        ["Total Drawings", str(total), "Production Ready", str(ready)],
        ["Needs Revision",  str(needs),  "Major Rework",    str(rework)],
        ["Avg Mfg. Score",  str(avg_sc), "", ""],
    ]
    st_table = Table(summary_data, colWidths=[42*mm, 20*mm, 42*mm, 20*mm])
    st_table.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0), (-1,-1), 9),
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (2,0), (2,-1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.HexColor('#fafafa'), colors.HexColor('#f3f3f3')]),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#e0e0e0')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(st_table)
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#eeeeee'), spaceAfter=4*mm))

    # Per-drawing entries
    story.append(Paragraph("DRAWING DETAILS", h2_s))
    for i, r in enumerate(results, 1):
        status = r.get("status","—")
        score  = r.get("manufacturability_score","—")
        clr    = '#16a34a' if "Ready" in status else ('#d97706' if "Revision" in status else '#dc2626')
        story.append(Paragraph(
            f'<font color="#f97316">{i}.</font>  <b>{r.get("drawing_name","—")}</b>  '
            f'<font color="{clr}">[ {status} ]</font>  '
            f'Score: <b>{score}/100</b>  |  Cost: <b>${r.get("estimated_cost_usd","—")}</b>  |  {r.get("complexity","—")} complexity',
            body_s
        ))
        story.append(Paragraph(r.get("summary","—"), ParagraphStyle('bs', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#555555'), leftIndent=12, spaceAfter=2)))

        issues = r.get("critical_issues",[])
        warns  = r.get("warnings",[])
        if issues:
            for iss in issues:
                story.append(Paragraph(f'<font color="#dc2626">  CRITICAL: {iss}</font>', ParagraphStyle('is', fontSize=8, fontName='Helvetica', leftIndent=12, spaceAfter=1)))
        if warns:
            for w in warns:
                story.append(Paragraph(f'<font color="#d97706">  WARNING: {w}</font>', ParagraphStyle('ws', fontSize=8, fontName='Helvetica', leftIndent=12, spaceAfter=1)))
        story.append(Spacer(1, 3*mm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dddddd')))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("Draft AI  ·  Powered by GPT-4o Vision", foot_s))

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_pdf(messages_display, drawing_name="drawing", title_block_data=None):
    _require_reportlab()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=20*mm, leftMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )

    title_style = ParagraphStyle(
        'T', fontSize=21, leading=24, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#f97316'), spaceAfter=0
    )
    sub_style = ParagraphStyle(
        'S', fontSize=12, leading=15, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#555555'), spaceAfter=0
    )
    meta_style = ParagraphStyle(
        'M', fontSize=8.5, leading=11, fontName='Helvetica',
        textColor=colors.HexColor('#888888')
    )
    q_style = ParagraphStyle(
        'Q', fontSize=10.5, leading=13, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#f97316'), spaceBefore=12, spaceAfter=5
    )
    a_style = ParagraphStyle(
        'A', fontSize=9.5, leading=15, fontName='Helvetica',
        textColor=colors.HexColor('#333333'), spaceAfter=6, leftIndent=10
    )
    tb_key = ParagraphStyle(
        'TK', fontSize=9, leading=12, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#444444')
    )
    tb_val = ParagraphStyle(
        'TV', fontSize=9, leading=12, fontName='Helvetica',
        textColor=colors.HexColor('#222222')
    )
    foot_style = ParagraphStyle(
        'F', fontSize=7.5, leading=10, fontName='Helvetica',
        textColor=colors.HexColor('#888888'), alignment=TA_CENTER
    )
    sec_style = ParagraphStyle(
        'SEC', fontSize=11, leading=14, fontName='Helvetica-Bold',
        textColor=colors.HexColor('#222222'), spaceBefore=10, spaceAfter=6
    )

    story = []

    # Header block — logo left, meta right, no overlap
    header_data = [[
        Paragraph('<font color="#f97316"><b>Draft AI</b></font>', ParagraphStyle('HL', fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#f97316'), leading=26)),
        Paragraph(
            f'<font color="#888888">Generated: {datetime.now().strftime("%d %B %Y, %I:%M %p")}</font><br/>'
            f'<font color="#aaaaaa">File: {drawing_name}</font>',
            ParagraphStyle('HR', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#888888'), leading=12, alignment=2)
        ),
    ]]
    header_tbl = Table(header_data, colWidths=[80*mm, None])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph("Engineering Drawing Analysis Report",
        ParagraphStyle('SUB', fontSize=11, fontName='Helvetica', textColor=colors.HexColor('#555555'), spaceAfter=0)))
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#f97316'), spaceAfter=6*mm))

    if title_block_data:
        story.append(Paragraph("TITLE BLOCK", sec_style))
        table_data = []
        for line in title_block_data.strip().split("\n"):
            if ":" in line:
                parts = line.split(":", 1)
                key = parts[0].strip()
                val = parts[1].strip() if len(parts) > 1 else ""
                if val and val.lower() != "not specified":
                    table_data.append([Paragraph(key, tb_key), Paragraph(val, tb_val)])
        if table_data:
            table = Table(table_data, colWidths=[50 * mm, 110 * mm])
            table.setStyle(TableStyle([
                ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#fafafa'), colors.HexColor('#f3f3f3')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(table)
            story.append(Spacer(1, 6 * mm))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#eeeeee'), spaceAfter=4 * mm))

    story.append(Paragraph("ANALYSIS", sec_style))
    q_num = 1
    i = 0
    while i < len(messages_display):
        msg = messages_display[i]
        if msg["role"] == "user":
            story.append(Paragraph(f"Q{q_num}: {msg['content']}", q_style))
            q_num += 1
            answer = ""
            if i + 1 < len(messages_display):
                answer = messages_display[i + 1]["content"]
                if answer.startswith("__TB__"):
                    answer = answer[6:]
                elif answer.startswith("__DIM__"):
                    answer = answer[7:]
            clean = answer.replace("**", "").replace("*", "")
            clean = clean.replace("\n", "<br/>")
            story.append(Paragraph(clean, a_style))
            story.append(HRFlowable(width="100%", thickness=0.4, color=colors.HexColor('#eeeeee'), spaceBefore=4 * mm, spaceAfter=2 * mm))
            i += 2
        else:
            i += 1

    story.append(Spacer(1, 7 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dddddd')))
    story.append(Spacer(1, 3.5 * mm))
    story.append(Paragraph("Made with Draft AI | Powered by GPT-4o Vision", foot_style))

    doc.build(story)
    buffer.seek(0)
    return buffer
# ══════════════════════════════════════════════════════════════════
# FEATURE: BOM GENERATOR
# ══════════════════════════════════════════════════════════════════

BOM_PROMPT = f"""You are a senior mechanical engineer and drawing interpreter specializing in assembly drawings and Bills of Materials.
{FORMAT_RULES}
Analyze this engineering drawing and extract every component or part visible.
Return ONLY a valid JSON object — no explanation, no markdown, no backticks.

Return exactly this structure:
{{
  "assembly_name": "Name from title block or inferred",
  "drawing_number": "From title block or Not specified",
  "revision": "From title block or Not specified",
  "date": "From title block or Not specified",
  "items": [
    {{
      "item_no": 1,
      "part_number": "PN-001 or Not specified",
      "description": "Hex Head Bolt",
      "quantity": 4,
      "material": "Steel AISI 1045 or Not specified",
      "standard": "ISO 4014 or Not specified",
      "finish": "Zinc plated or Not specified",
      "notes": "Any special callout or blank"
    }}
  ],
  "summary": "Brief assembly description and total item count"
}}

Rules:
- item_no must be sequential integers starting at 1
- quantity must be an integer
- If the drawing has a BOM table/parts list already visible, extract it exactly
- If it is a detail drawing with no BOM, analyze the part itself as a single item
- If quantity is unclear from the drawing, use 1
- Return ONLY the JSON. Nothing else."""


def generate_bom(image_file):
    """Extract BOM data from an assembly drawing. Returns parsed JSON dict."""
    result = _call_vision_api(
        image_file,
        BOM_PROMPT,
        "Extract the complete Bill of Materials from this engineering drawing. Return structured JSON only.",
        max_tokens=2000
    )
    import json, re
    clean = result.strip()
    if "```" in clean:
        clean = re.sub(r'```[a-z]*', '', clean).replace("```", "").strip()
    return json.loads(clean)


def generate_bom_excel(bom_data):
    """Generate a professional BOM Excel file from parsed BOM JSON."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill of Materials"

    ORANGE   = "F97316"
    DARK     = "0D0D0D"
    HEADER   = "1A1A1A"
    ROW_A    = "FAFAFA"
    ROW_B    = "FFFFFF"
    ORANGE_L = "FFF7ED"

    thin   = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"].value     = "BILL OF MATERIALS"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=ORANGE)
    ws["A1"].fill      = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    meta = [
        ("Assembly:", bom_data.get("assembly_name", "Not specified")),
        ("Drawing No.:", bom_data.get("drawing_number", "Not specified")),
        ("Revision:", bom_data.get("revision", "Not specified")),
        ("Date:", bom_data.get("date", "Not specified")),
        ("Generated:", datetime.now().strftime("%d %B %Y, %I:%M %p")),
    ]
    for r, (k, v) in enumerate(meta, 2):
        ws.cell(row=r, column=1, value=k).font = Font(name="Arial", size=9, bold=True, color="555555")
        ws.cell(row=r, column=2, value=v).font = Font(name="Arial", size=9, color="111111")
        ws.row_dimensions[r].height = 15

    headers    = ["ITEM", "PART NUMBER", "DESCRIPTION", "QTY", "MATERIAL", "STANDARD/SPEC", "FINISH", "NOTES"]
    col_widths = [6, 18, 32, 6, 24, 20, 18, 28]
    hrow = len(meta) + 2
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=hrow, column=ci, value=h)
        cell.font      = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hrow].height = 20

    items = bom_data.get("items", [])
    for ri, item in enumerate(items, 1):
        row = hrow + ri
        bg  = ROW_A if ri % 2 == 0 else ROW_B
        vals = [
            item.get("item_no", ri),
            item.get("part_number", "Not specified"),
            item.get("description", "Not specified"),
            item.get("quantity", 1),
            item.get("material", "Not specified"),
            item.get("standard", "Not specified"),
            item.get("finish", "Not specified"),
            item.get("notes", ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = border
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 4) else "left",
                vertical="center",
                wrap_text=(ci in (3, 5, 8))
            )
        ws.row_dimensions[row].height = 16

    # Totals row
    total_row = hrow + len(items) + 1
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws.cell(row=total_row, column=1, value=f"TOTAL ITEMS: {len(items)}").font = Font(name="Arial", size=9, bold=True, color=ORANGE)
    ws.cell(row=total_row, column=1).fill      = PatternFill("solid", fgColor=ORANGE_L)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    qty_formula = f"=SUM(D{hrow+1}:D{hrow+len(items)})"
    ws.cell(row=total_row, column=4, value=qty_formula).font = Font(name="Arial", size=9, bold=True)
    ws.cell(row=total_row, column=4).fill      = PatternFill("solid", fgColor=ORANGE_L)
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    for ci in range(5, 9):
        ws.cell(row=total_row, column=ci).fill = PatternFill("solid", fgColor=ORANGE_L)

    foot_row = total_row + 2
    ws.merge_cells(f"A{foot_row}:H{foot_row}")
    ws.cell(row=foot_row, column=1, value="Generated by Draft AI  |  Powered by GPT-4o Vision").font = Font(name="Arial", size=7, color="AAAAAA", italic=True)
    ws.cell(row=foot_row, column=1).alignment = Alignment(horizontal="center")



def generate_bom_pdf(bom_data):
    """Generate a clean PDF report from parsed BOM JSON."""
    _require_reportlab()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=18*mm, leftMargin=18*mm,
        topMargin=18*mm, bottomMargin=18*mm
    )
    title_s  = ParagraphStyle('BT',  fontSize=22, fontName='Helvetica-Bold', textColor=colors.HexColor('#f97316'), leading=26)
    sub_s    = ParagraphStyle('BS',  fontSize=10, fontName='Helvetica',      textColor=colors.HexColor('#555555'), spaceAfter=2)
    label_s  = ParagraphStyle('BL',  fontSize=8,  fontName='Helvetica-Bold', textColor=colors.HexColor('#888888'), leading=11, spaceAfter=0)
    meta_s   = ParagraphStyle('BM',  fontSize=8,  fontName='Helvetica',      textColor=colors.HexColor('#aaaaaa'), leading=11, alignment=2)
    th_s     = ParagraphStyle('BTH', fontSize=8,  fontName='Helvetica-Bold', textColor=colors.white,               leading=11)
    td_s     = ParagraphStyle('BTD', fontSize=8,  fontName='Helvetica',      textColor=colors.HexColor('#222222'), leading=11)
    foot_s   = ParagraphStyle('BF',  fontSize=7,  fontName='Helvetica',      textColor=colors.HexColor('#aaaaaa'), alignment=TA_CENTER)

    story = []
    items = bom_data.get("items", [])
    total_qty = sum(int(it.get("quantity", 1)) for it in items)

    # Header — no overlap
    hdr_data = [[
        Paragraph('<font color="#f97316"><b>Draft AI</b></font>', title_s),
        Paragraph(datetime.now().strftime("%d %B %Y, %I:%M %p"), meta_s),
    ]]
    hdr_tbl = Table(hdr_data, colWidths=[80*mm, None])
    hdr_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),0),('TOPPADDING',(0,0),(-1,-1),0)]))
    story.append(hdr_tbl)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph("Bill of Materials Report", sub_s))
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#f97316'), spaceAfter=4*mm))

    # Assembly metadata
    meta_rows = [
        ["Assembly", bom_data.get("assembly_name","—"),  "Drawing No.", bom_data.get("drawing_number","—")],
        ["Revision",  bom_data.get("revision","—"),      "Date",        bom_data.get("date","—")],
        ["Total Parts", str(len(items)),                 "Total Qty",   str(total_qty)],
    ]
    lbl_s2 = ParagraphStyle('L2', fontSize=8, fontName='Helvetica-Bold', textColor=colors.HexColor('#444444'))
    val_s2 = ParagraphStyle('V2', fontSize=8, fontName='Helvetica',      textColor=colors.HexColor('#111111'))
    meta_tbl_data = [[Paragraph(r[0],lbl_s2), Paragraph(str(r[1]),val_s2), Paragraph(r[2],lbl_s2), Paragraph(str(r[3]),val_s2)] for r in meta_rows]
    meta_tbl = Table(meta_tbl_data, colWidths=[32*mm, 55*mm, 32*mm, 55*mm])
    meta_tbl.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),'Helvetica'), ('FONTSIZE',(0,0),(-1,-1),8),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.HexColor('#fafafa'),colors.HexColor('#f3f3f3')]),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e0e0e0')),
        ('PADDING',(0,0),(-1,-1),5),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 5*mm))

    # BOM table
    headers = ["#", "Part No.", "Description", "Qty", "Material", "Standard", "Finish", "Notes"]
    col_w   = [10*mm, 28*mm, 44*mm, 10*mm, 30*mm, 24*mm, 22*mm, 6*mm]
    tbl_data = [[Paragraph(h, th_s) for h in headers]]
    for i, item in enumerate(items):
        row_s = ParagraphStyle(f'r{i}', fontSize=7.5, fontName='Helvetica',
                               textColor=colors.HexColor('#222222'), leading=10)
        tbl_data.append([
            Paragraph(str(item.get("item_no", i+1)), row_s),
            Paragraph(str(item.get("part_number","—")), row_s),
            Paragraph(str(item.get("description","—")), row_s),
            Paragraph(str(item.get("quantity",1)), row_s),
            Paragraph(str(item.get("material","—")), row_s),
            Paragraph(str(item.get("standard","—")), row_s),
            Paragraph(str(item.get("finish","—")), row_s),
            Paragraph(str(item.get("notes","")), row_s),
        ])

    bom_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    row_colors = []
    for ri in range(1, len(tbl_data)):
        bg = colors.HexColor('#fafafa') if ri % 2 == 0 else colors.white
        row_colors.append(('ROWBACKGROUNDS',(0,ri),(- 1,ri),[bg]))
    bom_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#e0e0e0')),
        ('PADDING',    (0,0), (-1,-1), 4),
        ('VALIGN',     (0,0), (-1,-1), 'TOP'),
        *row_colors,
    ]))
    story.append(bom_tbl)
    story.append(Spacer(1, 4*mm))

    # Summary note
    if bom_data.get("summary"):
        story.append(Paragraph(bom_data["summary"],
            ParagraphStyle('SUM', fontSize=8, fontName='Helvetica', textColor=colors.HexColor('#555555'), leading=12, leftIndent=4)))
        story.append(Spacer(1, 4*mm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#dddddd')))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph("Draft AI  ·  Bill of Materials  ·  Powered by GPT-4o Vision", foot_s))

    doc.build(story)
    buffer.seek(0)
    return buffer


# ══════════════════════════════════════════════════════════════════
# FEATURE: DRAWING STANDARDS CHECKER
# ══════════════════════════════════════════════════════════════════

STANDARDS_CHECKER_PROMPT = f"""You are a certified drawing checker with expertise in ASME Y14.5, ISO GPS, BS 8888, and industrial drawing standards.
{FORMAT_RULES}
Perform a comprehensive standards compliance check and return ONLY a valid JSON object — no explanation, no markdown, no backticks.

Return exactly this structure:
{{
  "overall_score": 78,
  "standard_detected": "ASME Y14.5 or ISO or BS 8888 or Unknown",
  "verdict": "PASS or CONDITIONAL PASS or FAIL",
  "checks": [
    {{
      "category": "Title Block",
      "status": "PASS",
      "score": 95,
      "findings": ["Part name present", "Drawing number present"],
      "violations": []
    }}
  ],
  "critical_violations": ["List of FAIL-level issues only"],
  "warnings": ["List of minor issues"],
  "recommendations": ["Ordered list of top fixes"],
  "summary": "One paragraph executive summary"
}}

Check ALL of the following categories:
1. Title Block: part name, number, revision, scale, date, drawn by, material, tolerances
2. Dimensioning: no duplicate dims, no missing dims, proper leader lines, dimension placement
3. Tolerancing: general tolerance stated, functional tolerances present, values realistic
4. GD&T Application: symbols correct per standard, datum references complete, FCF valid
5. Views and Projections: sufficient views, correct projection angle, section callouts correct
6. Annotations and Notes: surface finish symbols, thread callouts complete, weld symbols
7. Line Types and Weights: visible, hidden, center, dimension, section lines correct
8. Scale and Units: scale stated, consistent units throughout
Return ONLY the JSON. Nothing else."""


def check_drawing_standards(image_file):
    """Full standards compliance check. Returns parsed JSON dict."""
    result = _call_vision_api(
        image_file,
        STANDARDS_CHECKER_PROMPT,
        "Perform a complete drawing standards compliance check on this engineering drawing. Return structured JSON only.",
        max_tokens=2200
    )
    import json, re
    clean = result.strip()
    if "```" in clean:
        clean = re.sub(r'```[a-z]*', '', clean).replace("```", "").strip()
    return json.loads(clean)


# ══════════════════════════════════════════════════════════════════
# FEATURE: TEAM WORKSPACE
# ══════════════════════════════════════════════════════════════════

import uuid

WORKSPACE_FILE = "workspace.json"
WORKSPACE_DIR  = "workspace_drawings"


def _ensure_workspace_dir():
    Path(WORKSPACE_DIR).mkdir(exist_ok=True)


def load_workspace():
    if os.path.exists(WORKSPACE_FILE):
        try:
            with open(WORKSPACE_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {"projects": {}}


def save_workspace(ws):
    with open(WORKSPACE_FILE, "w") as f:
        json.dump(ws, f, indent=2)


def workspace_create_project(name, description="", created_by=""):
    _ensure_workspace_dir()
    ws  = load_workspace()
    pid = str(uuid.uuid4())[:8]
    ws["projects"][pid] = {
        "id":          pid,
        "name":        name,
        "description": description,
        "created_by":  created_by,
        "created_at":  datetime.now().isoformat(),
        "drawings":    {},
        "members":     [created_by] if created_by else [],
        "status":      "active",
    }
    save_workspace(ws)
    return pid


def workspace_add_drawing(project_id, uploaded_file, uploader, revision_note="Initial upload"):
    """Add a new drawing or new revision to a project. Returns drawing_id."""
    _ensure_workspace_dir()
    ws = load_workspace()
    if project_id not in ws["projects"]:
        raise ValueError("Project not found")

    proj = ws["projects"][project_id]
    name = uploaded_file.name
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    did  = f"{ts}_{name}"
    dest = os.path.join(WORKSPACE_DIR, did)

    uploaded_file.seek(0)
    raw = uploaded_file.read()
    with open(dest, "wb") as f:
        f.write(raw)
    uploaded_file.seek(0)

    existing = [d for d in proj["drawings"].values() if d["original_name"] == name]
    rev_num  = len(existing) + 1

    proj["drawings"][did] = {
        "id":            did,
        "original_name": name,
        "path":          dest,
        "revision":      rev_num,
        "uploaded_by":   uploader,
        "uploaded_at":   datetime.now().isoformat(),
        "revision_note": revision_note,
        "status":        "pending_review",
        "comments":      [],
        "size_mb":       round(len(raw) / (1024 * 1024), 2),
        "analysis":      None,
    }
    save_workspace(ws)
    return did


def workspace_add_comment(project_id, drawing_id, author, text, comment_type="comment"):
    """Add a comment or review action to a drawing."""
    ws   = load_workspace()
    proj = ws["projects"].get(project_id)
    if not proj:
        return
    drawing = proj["drawings"].get(drawing_id)
    if not drawing:
        return
    drawing["comments"].append({
        "id":     str(uuid.uuid4())[:8],
        "author": author,
        "text":   text,
        "type":   comment_type,
        "ts":     datetime.now().isoformat(),
    })
    if comment_type == "approval":
        drawing["status"] = "approved"
    elif comment_type == "rejection":
        drawing["status"] = "rejected"
    elif comment_type == "change_request":
        drawing["status"] = "needs_changes"
    save_workspace(ws)


def workspace_set_analysis(project_id, drawing_id, analysis_text):
    ws   = load_workspace()
    proj = ws["projects"].get(project_id)
    if proj and drawing_id in proj["drawings"]:
        proj["drawings"][drawing_id]["analysis"] = analysis_text
        save_workspace(ws)


def workspace_get_drawing_history(project_id, original_name):
    """Return all revisions of a filename, sorted oldest to newest."""
    ws   = load_workspace()
    proj = ws["projects"].get(project_id, {})
    revs = [d for d in proj.get("drawings", {}).values() if d["original_name"] == original_name]
    return sorted(revs, key=lambda d: d["revision"])


def workspace_delete_project(project_id):
    ws   = load_workspace()
    proj = ws["projects"].get(project_id)
    if proj:
        for d in proj["drawings"].values():
            try:
                os.remove(d["path"])
            except Exception:
                pass
        del ws["projects"][project_id]
        save_workspace(ws)